import os
import tkinter as tk
from tkinter import ttk
import yfinance as yf
import openpyxl
import pandas as pd
import ndjson


# メイン処理
def main():
    # 証券コード取得
    ticker_symbol = tab1_code_textbox.get()
    # コードに.T(東証)付ける
    t_code = addT(ticker_symbol)
    # コードから会社名取得
    com_name = getComInfo(t_code)['longName']
    # 最新日の株価データ取得
    ## 取得するカラム
    col_list = ["Open", "Close", "High", "Low"]
    ## 1日のデータ取得
    day_data = getDaydata(t_code)[col_list]
    # インデックス整形
    new_index_data = resetIndex(day_data,day_data.index)
    # 前日比カラム追加
    add_comparison_col = addComparisonCol(new_index_data)
    # 実行ファイルパス取得
    folder_path = getExcelFolderPath(com_name)
    file_path = folder_path + '\\data.xlsx'
    # excelへ保存
    writeDataToCsv(ticker_symbol, com_name, add_comparison_col)
    # 前日比算出
    get_file_data = valComparison(file_path)
    # 前日比未入力セルの削除
    delNotEnteredCompareCol(file_path)
    # 前日比入力済みデータを再入力
    saveEnteredCompareCol(file_path, get_file_data)
    
# コードに東証.Tをつける
def addT(code):
    return code + '.T'
    
# 会社データを取得
def getComInfo(code):
    try:
        com_info = yf.Ticker(code)
        return com_info.info
    except Exception as e:
        showInfo(e)

# 最新日の株価取得
def getDaydata(code):
    com = yf.Ticker(code)
    day_data = com.history(period="1d")
    return day_data

# インデックスの整形
def resetIndex(data_frame, index):
    old_index = pd.to_datetime(index)
    new_index = old_index.strftime('%m/%d')
    reset_index_data = data_frame.reset_index()
    drop_data_col = reset_index_data.drop(columns='Date')
    drop_data_col['new_date'] = new_index
    set_new_index = drop_data_col.set_index("new_date")
    return set_new_index

# 株価データを保存するときに重複をなくす
def avoidStockDataDuplication(file_path, csv_date):
    df = pd.read_excel(file_path, sheet_name="Sheet", header=None, skiprows=3)
    if (not df.empty):
        for date in df[0]:
            if (date == csv_date):
                # 一致する場合 True
                return True
            else:
                pass

# /stockまでのパスを返す
def getJsonFolderPath():
    # 実行フォルダパス取得
    exe_file_path = os.path.abspath(__file__)
    exe_folder_path = os.path.dirname(exe_file_path)
    stock_folder = exe_folder_path + '\\stock'
    return stock_folder

# /stock/com_name/data.xlsxパスを返す
def getExcelFolderPath(com_name):
    # 実行フォルダパス取得
    exe_file_path = os.path.abspath(__file__)
    exe_folder_path = os.path.dirname(exe_file_path)
    # stockフォルダ作成
    folder_path = exe_folder_path + '\\stock\\{}'.format(com_name)
    return folder_path

# エクセルにデータ書き込み
def writeDataToCsv(code, com_name, csv_data):
    # フォルダ作成
    folder_path = getExcelFolderPath(com_name)
    file_path = folder_path + '\\data.xlsx'
    if os.path.exists(folder_path) == False:
        # フォルダが存在しなかったら作成 
        os.makedirs(folder_path)
    # 事前にファイル,フォーマット作成
    if os.path.exists(file_path) == False:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "社名"
        ws["B1"] = com_name
        ws["A2"] = "証券コード"
        ws["B2"] = code
        ws["A3"] = "取得日"
        ws["B3"] = "始値"
        ws["C3"] = "前日比"
        ws["D3"] = "終値"
        ws["E3"] = "前日比"
        ws["F3"] = "高値"
        ws["G3"] = "前日比"
        ws["H3"] = "安値"
        ws["I3"] = "前日比"
        wb.save(file_path)
    # 重複チェック
    ## 重複がある場合は保存しない
    if (avoidStockDataDuplication(file_path, csv_data.index)):
        showInfo('既に保存されています')
    else:
        # xlsxファイルにデータ追記
        try:
            with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='overlay') as writer:
                # 最終行取得
                load_file = openpyxl.load_workbook(file_path)
                sheet = load_file['Sheet']
                max_row = sheet.max_row
                # 最終行に追記
                csv_data.to_excel(writer, sheet_name='Sheet', startrow=max_row, header=False)
            showInfo("データを取得しました")
        except Exception as e:
            showInfo(e)
        
# 前日比カラムの追加
def addComparisonCol(data):
    # 追加するカラム
    cols = ["Open", "Close", "High", "Low"]
    # カラムの位置
    col_num = 1
    for i in cols:
        data.insert(col_num, i + '前日比', "→")
        col_num += 2
    return data

# 前日比を割り出す
def valComparison(file_path):
    # 保存されているデータ全て取り出す
    df = pd.read_excel(file_path, sheet_name="Sheet", header=None, skiprows=3)
    # データ行数取得
    data_row = len(df)
    cols = [1, 3, 5, 7]
    start_comp_col = 2
    # 前日比算出
    for i in cols:
        for j in range(data_row):
            #　基準日と次の日の値取得
            ## 最終行
            if (j == data_row - 1):
                # 基準日
                day_val = df.iat[j, i]
                # 次の日
                next_day_val = df.iat[j, i]
                # 前日比入力
                if (day_val > next_day_val):
                    df.iat[j, start_comp_col] = "↓"
                elif (day_val < next_day_val):
                    df.iat[j, start_comp_col] = "↑"
            else:
                # 基準日
                day_val = df.iat[j, i]
                # 次の日
                next_day_val = df.iat[j + 1, i]
                # 前日比入力
                if (day_val > next_day_val):
                    # >の場合
                    df.iat[j + 1, start_comp_col] = "↓"
                elif (day_val < next_day_val):
                    # <の場合
                    df.iat[j + 1, start_comp_col] = "↑"
        start_comp_col += 2
    return df

# 前日比未入力データを削除
def delNotEnteredCompareCol(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb['Sheet']
    # データの最終行を取得
    max_data_row = ws.max_row
    # データ行を削除
    ws.delete_rows(4, max_data_row)
    # 保存
    wb.save(file_path)

# 前日比カラムが入力済みのデータを保存
def saveEnteredCompareCol(file_path, data):
    with pd.ExcelWriter(file_path, mode="a", if_sheet_exists="overlay") as writer:
        data.to_excel(writer, sheet_name="Sheet", index=False, header=False, startrow=3)

# インフォメーションウィンドウ表示
def showInfo(error_messages):
    # サブウィンドウ
    sub_window = tk.Toplevel()
    sub_window.geometry('300x300')
    # エラーメッセージ表示
    sub_error_message = tk.Label(sub_window, text=error_messages)
    sub_error_message.place(x=30, y=70)
    # サブウィンドウ閉じる
    def closeWindow():
        sub_window.destroy()
    # 閉じるボタン表示
    sub_button = tk.Button(master=sub_window, text="閉じる", command=closeWindow)
    sub_button.place(x=30, y=100)

# Jsonデータの重複チェック
## 重複あり True, 重複なし False
def jsonAvoidDuplication(file_path, key):
    result = False
    with open(file_path, mode='r') as f:
        code_set_data = ndjson.load(f)
        for i in code_set_data:
            for key_code in i.keys():
                if (key_code == key):
                    result = True
                    return result
    return result

# 証券コード保存
def saveCodeList():
    # コードから社名を取得
    # 証券コード取得
    ticker_symbol = tab2_code_textbox.get()
    # コードに.T(東証)付ける
    t_code = addT(ticker_symbol)
    # コードから会社名取得
    com_name = getComInfo(t_code)['longName']
    # 実行ファイル配下にフォルダ作成
    folder_path = getJsonFolderPath()
    #folder_path = 
    # フォルダの存在確認
    if os.path.exists(folder_path) == False:
        # フォルダが存在しなかったら作成 
        os.makedirs(folder_path)
    # ファイルパス
    file_path = folder_path + '\\code_set.json'
    # JSON data
    ## key = 証券コード, value = 社名
    code_set = {ticker_symbol:com_name}
    # 重複チェック
    ## True(重複あり)の時にメッセージ表示
    if (jsonAvoidDuplication(file_path, ticker_symbol)):
        showInfo('既に保存されているコードです')
    else:
        # False(重複なし)
        # データに追加
        with open(file_path, mode='a') as f:    
            writer = ndjson.writer(f)
            writer.writerow(code_set)
        # 保存出来たらサブウィンドウでメッセージ表示
        showInfo('保存しました')
    
# コードリストから株価を取得する
def getDataFromCodeList():
    # JSONファイルパス取得
    json_folder_path = getJsonFolderPath()
    json_file_path = json_folder_path + '\\code_set.json'
    with open(json_file_path, mode='r') as f:
        code_set_data = ndjson.load(f)
        for i in code_set_data:
            for key in i.keys():
                # 東証用に.Tをつける
                t_code = addT(key)
                # コードから会社名取得
                com_name = getComInfo(t_code)['longName']
                # 最新日の株価データ取得
                ## 取得するカラム
                col_list = ["Open", "Close", "High", "Low"]
                ## 1日のデータ取得
                day_data = getDaydata(t_code)[col_list]
                # インデックス整形
                new_index_data = resetIndex(day_data,day_data.index)
                # 前日比カラム追加
                add_comparison_col = addComparisonCol(new_index_data)
                # 実行ファイルパス取得
                folder_path = getExcelFolderPath(com_name)
                file_path = folder_path + '\\data.xlsx'
                # excelへ保存
                writeDataToCsv(key, com_name, add_comparison_col)
                # 前日比算出
                get_file_data = valComparison(file_path)
                # 前日比未入力セルの削除
                delNotEnteredCompareCol(file_path)
                # 前日比入力済みデータを再入力
                saveEnteredCompareCol(file_path, get_file_data)
                

# JSONファイルに保存されているコードを読み込み、表示する
def showCodeSetTable(window):
    # 証券コードが保存されていたら表示
    columns = ('証券コード', '会社名')
    # tableの作成
    tree = ttk.Treeview(window, columns=columns)
    tree.heading('証券コード', text='証券コード')
    tree.heading('会社名', text='会社名')
    try:        
        # jsonファイルパス
        folder_path = getJsonFolderPath()
        json_file_path = folder_path + '\\code_set.json'
        # JSONデータの読み込み
        with open(json_file_path, mode='r') as f:
            code_set_data = ndjson.load(f)
        for i in code_set_data:
            for k, v in i.items():
                # tableに値の挿入
                tree.insert(parent='', index='end', values=(k, v))
    except FileNotFoundError:
        return tk.Label(master=window, text='コードが保存されていません')
    except Exception as e:
        print(e)
    return tree

# メインウィンドウを作成
baseGround = tk.Tk()
# ウィンドウのサイズを設定
baseGround.geometry('600x600')
# 画面タイトル
baseGround.title('株式データ取得')
# Notebookウィジェットの作成
notebook = ttk.Notebook(baseGround)
# タブの作成
tab_one = tk.Frame(notebook, bg='white')
tab_two = tk.Frame(notebook, bg='white')
# notebookにタブを追加
notebook.add(tab_one, text='個別取得')
notebook.add(tab_two, text='まとめて取得')
# tabの配置
notebook.pack(expand=True, fill='both', padx=10, pady=10)
# tab1 ラベル
tab1_code_label = tk.Label(tab_one, text='証券コード')
tab1_code_label.place(x=30, y=70)
# tab1 テキストボックス
tab1_code_textbox = tk.Entry(tab_one, width=40)
tab1_code_textbox.place(x=30, y=90)
# tab1 取得ボタン
tab1_button = tk.Button(tab_one, text='取得', command=main).place(x=30, y=120)
# tab2 ラベル
tab2_code_label = tk.Label(tab_two, text='保存する証券コード')
tab2_code_label.place(x=30, y=70)
# tab2 テキストボックス
tab2_code_textbox = tk.Entry(tab_two, width=40)
tab2_code_textbox.place(x=30, y=90)
# tab2 保存ボタン
tab2_save_button = tk.Button(tab_two, text='保存ボタン', command=saveCodeList).place(x=30, y=120)
# tab2 取得ボタン
tab2_button = tk.Button(tab_two, text='取得', command=getDataFromCodeList).place(x=30, y=150)
# tab2 保存済みコード表示テキストラベル
set_code_label = tk.Label(tab_two, text='保存済みのコード')
set_code_label.place(x=30, y=200)
# tab2 JSONファイルのデータをtableで表示
code_set_table = showCodeSetTable(tab_two)
code_set_table.place(x=30, y=230)

baseGround.mainloop()